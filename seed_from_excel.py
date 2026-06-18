"""
One-time seed script: imports all existing data from the client's Excel files
into the EventBot database so the tracker starts pre-populated.
Run once: python seed_from_excel.py
"""

import os
import sys
import re
import logging
from datetime import datetime

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))
from deduplicator import upsert_records

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

ATTACHMENTS = os.path.join(
    os.path.dirname(__file__), "..", "workrbee-attachments"
)

# Map each Excel file + sheet to venue metadata
SHEET_VENUE_MAP = {
    "DC and Gaylord   2021 - 2023 (1).xlsx": {
        "Gaylord": {"venue_name": "Gaylord National Harbor", "city": "National Harbor", "state": "MD"},
        "DC Convention Center": {"venue_name": "DC Convention Center", "city": "Washington", "state": "DC"},
        "Marriott Marquis": {"venue_name": "Marriott Marquis", "city": "Washington", "state": "DC"},
        "Hilton": {"venue_name": "Hilton Washington DC Capitol Hill", "city": "Washington", "state": "DC"},
        "Renaissance": {"venue_name": "Renaissance Washington DC", "city": "Washington", "state": "DC"},
        "Grand Hyatt Washington": {"venue_name": "Grand Hyatt Washington", "city": "Washington", "state": "DC"},
        "Omni Shoreham Hotel": {"venue_name": "Omni Shoreham Hotel", "city": "Washington", "state": "DC"},
        "Omni Shoreham Hotel ": {"venue_name": "Omni Shoreham Hotel", "city": "Washington", "state": "DC"},
        "JW Marriott": {"venue_name": "JW Marriott Washington DC", "city": "Washington", "state": "DC"},
    },
    "DC and Gaylord   2021 - 2023 copy.xlsx": {
        "Gaylord": {"venue_name": "Gaylord National Harbor", "city": "National Harbor", "state": "MD"},
    },
    "Gaylord Events 2025 - 2-26.xlsx": {
        "Gaylord": {"venue_name": "Gaylord National Harbor", "city": "National Harbor", "state": "MD"},
        "Bethesda": {"venue_name": "Bethesda North Marriott Hotel & Conference Center", "city": "Bethesda", "state": "MD"},
        "Hyatt": {"venue_name": "Hyatt Regency Bethesda", "city": "Bethesda", "state": "MD"},
        "Bethesdan": {"venue_name": "The Bethesdan Hotel", "city": "Bethesda", "state": "MD"},
        "Harborside": {"venue_name": "Harborside Hotel National Harbor", "city": "National Harbor", "state": "MD"},
        "MGM": {"venue_name": "MGM National Harbor", "city": "National Harbor", "state": "MD"},
    },
    "Baltimore Hotels 2026.xlsx": {
        "Baltimore Convention Center": {"venue_name": "Baltimore Convention Center", "city": "Baltimore", "state": "MD"},
        "Hilton": {"venue_name": "Hilton Baltimore Inner Harbor", "city": "Baltimore", "state": "MD"},
        "Marriott IH": {"venue_name": "Marriott Inner Harbor at Camden Yards", "city": "Baltimore", "state": "MD"},
        "Four Seasons": {"venue_name": "Four Seasons Baltimore", "city": "Baltimore", "state": "MD"},
        "Embassy Suites": {"venue_name": "Embassy Suites Baltimore Inner Harbor", "city": "Baltimore", "state": "MD"},
        "Hyatt Regency IH": {"venue_name": "Hyatt Regency Baltimore Inner Harbor", "city": "Baltimore", "state": "MD"},
        "Marriott Waterfront": {"venue_name": "Baltimore Marriott Waterfront", "city": "Baltimore", "state": "MD"},
        "Renaissance BHH": {"venue_name": "Renaissance Baltimore Harborplace Hotel", "city": "Baltimore", "state": "MD"},
        "Courtyard Marriott": {"venue_name": "Courtyard Baltimore Downtown Inner Harbor", "city": "Baltimore", "state": "MD"},
        "Hilton Garden Inn": {"venue_name": "Hilton Garden Inn Baltimore Inner Harbor", "city": "Baltimore", "state": "MD"},
        "Renaissance Harborplace": {"venue_name": "Renaissance Harborplace Hotel Baltimore", "city": "Baltimore", "state": "MD"},
        "Hotel Indigo": {"venue_name": "Hotel Indigo Baltimore Downtown", "city": "Baltimore", "state": "MD"},
        "Lord Baltimore": {"venue_name": "Lord Baltimore Hotel", "city": "Baltimore", "state": "MD"},
        "Hampton Inn": {"venue_name": "Hampton Inn Baltimore Convention Center", "city": "Baltimore", "state": "MD"},
        "Hampton Inn ": {"venue_name": "Hampton Inn Baltimore Convention Center", "city": "Baltimore", "state": "MD"},
    },
    "Delaware Event List.xlsx": {
        "Chase Center on the Riverfront": {"venue_name": "Chase Center on the Riverfront", "city": "Wilmington", "state": "DE"},
        "Hotel DuPont": {"venue_name": "Hotel DuPont", "city": "Wilmington", "state": "DE"},
        "Hotel DuPont ": {"venue_name": "Hotel DuPont", "city": "Wilmington", "state": "DE"},
        "DoubleTree by Hilton": {"venue_name": "DoubleTree by Hilton Wilmington", "city": "Wilmington", "state": "DE"},
    },
    "Philadelphia Event List - 2025.xlsx": {
        "Philly Convention Center": {"venue_name": "Pennsylvania Convention Center", "city": "Philadelphia", "state": "PA"},
        "Marriott": {"venue_name": "Philadelphia Marriott Downtown", "city": "Philadelphia", "state": "PA"},
        "Lowes Hotel": {"venue_name": "Loews Philadelphia Hotel", "city": "Philadelphia", "state": "PA"},
        "Element Philadelphia": {"venue_name": "Element Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Hyatt CC": {"venue_name": "Hyatt Centric Center City Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Sheraton Philadelphia Downtown": {"venue_name": "Sheraton Philadelphia Downtown", "city": "Philadelphia", "state": "PA"},
        "W Philadelphia": {"venue_name": "W Philadelphia", "city": "Philadelphia", "state": "PA"},
        "The Bellevue": {"venue_name": "The Bellevue Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "The Bellevue ": {"venue_name": "The Bellevue Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Holiday - Drexel Hill": {"venue_name": "Holiday Inn Drexel Hill Philadelphia", "city": "Drexel Hill", "state": "PA"},
        "Temple University ": {"venue_name": "Temple University Conference Center", "city": "Philadelphia", "state": "PA"},
        "Circa Center": {"venue_name": "Circa Centre Philadelphia", "city": "Philadelphia", "state": "PA"},
        "The Logan": {"venue_name": "The Logan Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Four Seasons": {"venue_name": "Four Seasons Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Sofitel": {"venue_name": "Sofitel Philadelphia at Rittenhouse Square", "city": "Philadelphia", "state": "PA"},
        "The Ritz-Carlton": {"venue_name": "The Ritz-Carlton Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Union League": {"venue_name": "The Union League of Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Sonesta": {"venue_name": "Sonesta Philadelphia Downtown Rittenhouse Square", "city": "Philadelphia", "state": "PA"},
        "Marriott Old City": {"venue_name": "Marriott Philadelphia Old City", "city": "Philadelphia", "state": "PA"},
        "Kimpton - Palamor": {"venue_name": "Kimpton Palomar Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Kimpton-Monaco": {"venue_name": "Kimpton Monaco Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Live Casino": {"venue_name": "Live Casino & Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Hilton Penns Landing": {"venue_name": "Hilton Philadelphia Penn's Landing", "city": "Philadelphia", "state": "PA"},
        "Airport Marriott": {"venue_name": "Airport Marriott Philadelphia", "city": "Philadelphia", "state": "PA"},
        "VF Casino": {"venue_name": "Valley Forge Casino Resort", "city": "King of Prussia", "state": "PA"},
        "Oaks Expo Center": {"venue_name": "Oaks Expo Center", "city": "Oaks", "state": "PA"},
        "Temple University": {"venue_name": "Temple University Conference Services", "city": "Philadelphia", "state": "PA"},
        "Warwick Hotel": {"venue_name": "Warwick Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Notary Hotel": {"venue_name": "Notary Hotel Philadelphia", "city": "Philadelphia", "state": "PA"},
        "Others Philadelphia": {"venue_name": "Pennsylvania Convention Center", "city": "Philadelphia", "state": "PA"},
    },
}

# Sheets to skip entirely
SKIP_SHEETS = {
    "Wardman - DO NOT USE",
    "Washington DC Annual Events",
    "Annual Events",
    "Event websites",
    "Addition search",
    "Rita",
    "Hyatt Place IH",
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
PHONE_RE = re.compile(r"(\(?\d{3}\)?[\s\-.])\d{3}[\s\-.]\d{4}")


def _is_data_row(row_values: list) -> bool:
    """Return True if this row looks like an event record (not a header/label row)."""
    first = str(row_values[0]).strip() if row_values else ""
    # Skip blank rows, venue address rows, instruction rows, column headers
    skip_exact = {
        "name of meeting, convention or tradeshow",
        "name of meeting", "contact person", "email", "telephone",
        "date of the event", "e-mail sent", "call notes",
        "add events in chronological order", "add events in chronological by date order",
        "none", "nan", "", "date",
    }
    skip_starters = (
        "name of meeting", "name of convention",
        "add events", "staci will", "do not email",
        "meeting space", "http", "www.",
    )
    fl = first.lower().strip()
    if fl in skip_exact:
        return False
    if fl.startswith(skip_starters):
        return False
    # Skip month-label rows like "SEPTEMBER", "OCTOBER 2026"
    if re.match(r"^(january|february|march|april|may|june|july|august|september|october|november|december)\b", fl):
        return False
    # Skip address rows (contain "St," or "Ave," or "Blvd")
    if re.search(r"\b(st\.?,|ave\.?,|blvd\.?,|drive,|road,|suite)", fl):
        return False
    if len(first) < 5:
        return False
    return True


def parse_sheet(ws, venue_meta: dict) -> list[dict]:
    """Extract event records from one worksheet."""
    records = []
    rows = list(ws.iter_rows(values_only=True))

    for row in rows:
        vals = [str(v).strip() if v is not None else "" for v in row]
        if not any(vals):
            continue
        if not _is_data_row(vals):
            continue

        event_name = vals[0] if len(vals) > 0 else ""
        if len(event_name) < 4:
            continue

        contact = vals[1] if len(vals) > 1 else ""
        # Some sheets have email before phone, some phone before email — detect
        col2 = vals[2] if len(vals) > 2 else ""
        col3 = vals[3] if len(vals) > 3 else ""
        date_col = vals[4] if len(vals) > 4 else ""

        if "@" in col2:
            email = col2
            phone = col3
        elif "@" in col3:
            email = col3
            phone = col2
        else:
            # Try regex scan across all columns
            all_text = " ".join(vals)
            email_m = EMAIL_RE.search(all_text)
            phone_m = PHONE_RE.search(all_text)
            email = email_m.group() if email_m else ""
            phone = phone_m.group() if phone_m else ""

        email_sent = vals[5] if len(vals) > 5 else ""
        call1 = vals[6] if len(vals) > 6 else ""
        call2 = vals[7] if len(vals) > 7 else ""
        call3 = vals[8] if len(vals) > 8 else ""
        call4 = vals[9] if len(vals) > 9 else ""

        # Determine status from notes
        status = "New"
        notes_combined = " ".join([email_sent, call1, call2, call3, call4]).lower()
        if "booked" in notes_combined or "contract" in notes_combined:
            status = "Booked"
        elif "voicemail" in notes_combined or "vm" in notes_combined:
            status = "Voicemail"
        elif "call" in notes_combined and len(call1) > 3:
            status = "Called"
        elif email_sent and email_sent.lower() not in ("", "nan", "no"):
            status = "Emailed"

        records.append({
            "venue_name": venue_meta["venue_name"],
            "city": venue_meta["city"],
            "state": venue_meta["state"],
            "event_name": event_name,
            "event_dates": date_col,
            "contact_person": contact,
            "contact_title": "",
            "email": email.lower() if email else "",
            "phone": phone,
            "event_url": "",
            "email_sent": email_sent if email_sent.lower() not in ("nan", "") else "",
            "call_notes_1": call1 if call1.lower() != "nan" else "",
            "call_notes_2": call2 if call2.lower() != "nan" else "",
            "call_notes_3": call3 if call3.lower() != "nan" else "",
            "call_notes_4": call4 if call4.lower() != "nan" else "",
            "status": status,
            "scraped_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        })
    return records


def seed() -> None:
    all_records = []
    total_files = 0
    total_sheets = 0

    for filename, sheet_map in SHEET_VENUE_MAP.items():
        filepath = os.path.join(ATTACHMENTS, filename)
        if not os.path.exists(filepath):
            logger.warning("File not found: %s — skipping", filename)
            continue

        logger.info("Reading: %s", filename)
        total_files += 1

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("  Could not open %s: %s", filename, exc)
            continue

        for sheet_name in wb.sheetnames:
            # Strip trailing/leading spaces for comparison (Excel often adds them)
            sheet_name_stripped = sheet_name.strip()
            if sheet_name_stripped in SKIP_SHEETS:
                logger.info("  Skipping sheet: %s", sheet_name)
                continue

            # Try exact match first, then stripped match
            venue_meta = sheet_map.get(sheet_name) or sheet_map.get(sheet_name_stripped)
            if not venue_meta:
                logger.info("  No venue mapping for sheet: %s — skipping", sheet_name)
                continue

            ws = wb[sheet_name]
            records = parse_sheet(ws, venue_meta)
            logger.info("  Sheet %-40s → %d records", sheet_name, len(records))
            all_records.extend(records)
            total_sheets += 1

        wb.close()

    logger.info("")
    logger.info("Total records parsed: %d", len(all_records))

    if all_records:
        added, updated = upsert_records(all_records)
        logger.info("Seeded into DB: +%d new, ~%d updated", added, updated)
    else:
        logger.info("No records to seed.")

    logger.info("")
    logger.info("Seed complete! Files: %d | Sheets: %d | Records: %d",
                total_files, total_sheets, len(all_records))


if __name__ == "__main__":
    seed()
