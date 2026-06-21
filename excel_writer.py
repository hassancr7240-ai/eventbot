"""
Excel output writer — exactly matches the client's format.

Client format (from DC and Gaylord 2021-2023.xlsx):
Row 1: blank
Row 2: Venue name + address (merged)
Row 3: Date range (merged)
Row 4: Column headers
Row 5: "Add events in chronological order"
Row 6+: Data rows — sorted chronologically

Exact columns (matching client's Gaylord sheet):
  A: Name of Meeting, Convention or Tradeshow
  B: CONTACT PERSON
  C: EMAIL
  D: TELEPHONE
  E: DATE OF THE EVENT
  F: e-mail sent
  G-Q: Call Notes (11 columns, labelled Call Notes 1 through Call Notes 11)
"""

import os
import re
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from venues import VENUES
from deduplicator import load_db

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "output")

# ── EXACT COLUMN STRUCTURE matching client Excel ──────────────────────────────
HEADERS = [
    "Name of Meeting, Convention or Tradeshow",
    "CONTACT PERSON",
    "EMAIL",
    "TELEPHONE",
    "DATE OF THE EVENT",
    "e-mail sent",
    "Call Notes 1",
    "Call Notes 2",
    "Call Notes 3",
    "Call Notes 4",
    "Call Notes 5",
    "Call Notes 6",
    "Call Notes 7",
    "Call Notes 8",
    "Call Notes 9",
    "Call Notes 10",
    "Call Notes 11",
]

# Column widths matching client format
COL_WIDTHS = [42, 22, 30, 18, 22, 14, 28, 28, 28, 28, 28, 28, 28, 28, 28, 28, 28]

# ── COLOURS ───────────────────────────────────────────────────────────────────
NAVY      = "0F1F3D"
HDR_BG    = "1E3A5F"
WHITE     = "FFFFFF"
GRAY_ROW  = "F8FAFC"
BLUE_L    = "DBEAFE"
AMBER_L   = "FEF3C7"
PURPLE_L  = "EDE9FE"
GREEN_L   = "D1FAE5"
RED_L     = "FEE2E2"

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "new":        PatternFill("solid", fgColor=BLUE_L),
    "email":      PatternFill("solid", fgColor=AMBER_L),
    "call":       PatternFill("solid", fgColor=PURPLE_L),
    "voicemail":  PatternFill("solid", fgColor=PURPLE_L),
    "book":       PatternFill("solid", fgColor=GREEN_L),
    "contract":   PatternFill("solid", fgColor=GREEN_L),
    "do not":     PatternFill("solid", fgColor=RED_L),
}

def _row_fill(status: str) -> PatternFill:
    s = status.lower()
    for key, fill in STATUS_FILL.items():
        if key in s:
            return fill
    return PatternFill("solid", fgColor=WHITE)


# ── SORT KEY ──────────────────────────────────────────────────────────────────
_MONTH_NUM = {m.lower(): i+1 for i, m in enumerate([
    "january","february","march","april","may","june",
    "july","august","september","october","november","december"
])}

def _sort_key(rec: dict) -> tuple:
    date_str = (rec.get("event_dates") or "").lower()
    year, month, day = 9999, 99, 99
    ym = re.search(r"20(\d{2})", date_str)
    if ym:
        year = int("20" + ym.group(1))
    for name, num in _MONTH_NUM.items():
        if name in date_str:
            month = num
            break
    dm = re.search(r"\b(\d{1,2})\b", date_str)
    if dm:
        day = int(dm.group(1))
    return (year, month, day)


# ── VENUE LOOKUP ──────────────────────────────────────────────────────────────
def _venue_info(venue_name: str) -> dict | None:
    for v in VENUES:
        if v["name"] == venue_name:
            return v
    return None


# ── SHEET WRITER ──────────────────────────────────────────────────────────────
def _write_sheet(wb, venue_name: str, records: list[dict]) -> None:
    info = _venue_info(venue_name)
    address = info["address"] if info else ""
    city_state = f"{info['city']}, {info['state']}" if info else ""

    # Safe sheet name (Excel max 31 chars, no special chars)
    safe = re.sub(r'[\\/*?\[\]:]', '', venue_name)[:31]
    ws = wb.create_sheet(title=safe)

    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    # ── ROW 1: blank ──
    ws.row_dimensions[1].height = 6

    # ── ROW 2: venue name + address ──
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"{venue_name}    {address}"
    c.font = Font(bold=True, size=13, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 26

    # ── ROW 3: date range ──
    ws.merge_cells(f"A3:{last_col}3")
    c = ws["A3"]
    c.value = f"Events: {datetime.now().year} forward — updated automatically by EventBot"
    c.font = Font(italic=True, size=9, color="6B7280")
    c.fill = PatternFill("solid", fgColor="F9FAFB")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 15

    # ── ROW 4: column headers ──
    for col_idx, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=4, column=col_idx, value=header)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = "A5"

    # ── ROW 5: instruction ──
    ws.merge_cells(f"A5:{last_col}5")
    c = ws["A5"]
    c.value = "Add events in chronological order"
    c.font = Font(italic=True, size=9, color="9CA3AF")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[5].height = 14

    # ── DATA ROWS from row 6 ──
    sorted_records = sorted(records, key=_sort_key)

    for row_idx, rec in enumerate(sorted_records, start=6):
        status = rec.get("status", "New")
        fill = _row_fill(status)

        row_data = [
            rec.get("event_name", ""),
            _format_contact(rec),          # "Name, Title" combined like client format
            rec.get("email", ""),
            rec.get("phone", ""),
            rec.get("event_dates", ""),
            rec.get("email_sent", ""),
            rec.get("call_notes_1", ""),
            rec.get("call_notes_2", ""),
            rec.get("call_notes_3", ""),
            rec.get("call_notes_4", ""),
            "", "", "", "", "", "", "",     # call notes 5-11 blank, ready to fill
        ]

        for col_idx, value in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = Font(size=10)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(col_idx >= 7))
            c.fill = fill

            # Email as hyperlink
            if col_idx == 3 and value and "@" in str(value):
                c.hyperlink = f"mailto:{value}"
                c.font = Font(size=10, color="1D4ED8", underline="single")

        ws.row_dimensions[row_idx].height = 18

    # Auto-filter on header row
    ws.auto_filter.ref = f"A4:{last_col}4"

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_contact(rec: dict) -> str:
    """Combine name + title into 'Joan Smith, Conference Planner' format."""
    name = (rec.get("contact_person") or "").strip()
    title = (rec.get("contact_title") or "").strip()
    if name and title:
        return f"{name}, {title}"
    return name or title


# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
def _write_summary(wb, db: dict) -> None:
    ws = wb.create_sheet(title="SUMMARY", index=0)

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "EventBot Pro — Master Summary"
    c.font = Font(bold=True, size=16, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.now().strftime('%B %d, %Y  %I:%M %p')}"
    c.font = Font(italic=True, size=10, color="6B7280")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # Headers
    for col, label in enumerate(["Venue", "City", "State", "Events", "Contacts", "Emailed"], start=1):
        c = ws.cell(row=4, column=col, value=label)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=HDR_BG)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
    ws.row_dimensions[4].height = 24

    row = 5
    tot_events = tot_contacts = tot_emailed = 0

    for venue_name, records in sorted(db.items()):
        info = _venue_info(venue_name)
        city  = info["city"]  if info else ""
        state = info["state"] if info else ""
        events   = len({rec.get("event_name","") for rec in records})
        contacts = sum(1 for r in records if r.get("email"))
        emailed  = sum(1 for r in records if "yes" in str(r.get("email_sent","")).lower()
                       or "email" in str(r.get("status","")).lower())
        tot_events += events; tot_contacts += contacts; tot_emailed += emailed

        fill = PatternFill("solid", fgColor="F9FAFB" if row%2==0 else WHITE)
        for col, val in enumerate([venue_name, city, state, events, contacts, emailed], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(size=10)
            c.fill = fill
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col<=3 else "center")
        row += 1

    # Totals
    for col, val in enumerate(["TOTAL", "", "", tot_events, tot_contacts, tot_emailed], start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, size=11, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 22

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8
    for col in "DEF":
        ws.column_dimensions[col].width = 12


# ── PUBLIC ENTRY POINT ────────────────────────────────────────────────────────
def build_excel(filename: str | None = None) -> str:
    db = load_db()
    if not db:
        raise ValueError("No data in database. Run the bot or seed first.")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not filename:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"EventBot_Tracker_{ts}.xlsx"

    path = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_summary(wb, db)

    # Group by city_group so tabs are organised logically
    city_order = ["Washington DC", "National Harbor MD", "Bethesda MD",
                  "Baltimore MD", "Philadelphia PA", "Delaware"]
    grouped: dict[str, list[str]] = defaultdict(list)
    for vname in db:
        info = _venue_info(vname)
        cg = info["city_group"] if info else "Other"
        grouped[cg].append(vname)

    for cg in city_order + [k for k in grouped if k not in city_order]:
        for vname in grouped.get(cg, []):
            if db.get(vname):
                _write_sheet(wb, vname, db[vname])

    wb.save(path)
    return path


def get_latest_excel() -> str | None:
    if not os.path.exists(OUTPUT_DIR):
        return None
    files = sorted(
        [f for f in os.listdir(OUTPUT_DIR) if f.startswith("EventBot_Tracker") and f.endswith(".xlsx")],
        reverse=True,
    )
    return os.path.join(OUTPUT_DIR, files[0]) if files else None
