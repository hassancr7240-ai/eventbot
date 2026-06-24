"""
EventBot Pro - FINAL VERSION
Fresh start each search. No junk. No duplicates. Real contacts only.
Fast, clean, simple.
"""

import streamlit as st
import os
import json
import time
import subprocess
from datetime import datetime
import pandas as pd
from venues import VENUES

st.set_page_config(page_title="EventBot Pro", layout="wide", initial_sidebar_state="collapsed")

# ═══════════════════════════════════════════════════════════════════════════════
# THEME
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
body { background: #f5f5f5; font-family: 'Segoe UI', sans-serif; }
.main { background: #ffffff; }
h1 { color: #1a1a1a; margin-bottom: 0.5rem; }
h2 { color: #2d2d2d; margin-top: 1rem; }

.stat-box {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white; padding: 25px; border-radius: 10px; text-align: center;
}
.stat-num { font-size: 2.2rem; font-weight: bold; }
.stat-label { font-size: 0.85rem; opacity: 0.9; margin-top: 8px; }

.status-running { background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 5px; }
.status-done { background: #d4edda; border-left: 4px solid #28a745; padding: 15px; border-radius: 5px; }
.status-error { background: #f8d7da; border-left: 4px solid #dc3545; padding: 15px; border-radius: 5px; }

.stButton > button { background: linear-gradient(135deg, #667eea, #764ba2); color: white; border: none; width: 100%; }
.stButton > button:hover { opacity: 0.9; }

</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    st.markdown("<h1 style='text-align: center; margin-top: 15rem;'>🔒 EventBot Pro</h1>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pwd = st.text_input("Password", type="password")
        if st.button("Login", use_container_width=True):
            if pwd == "workrbee2026":
                st.session_state.auth = True
                st.rerun()
            else:
                st.error("❌ Wrong password")
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════════
# SESSION
# ═══════════════════════════════════════════════════════════════════════════════

if "current_results" not in st.session_state:
    st.session_state.current_results = []
if "search_running" not in st.session_state:
    st.session_state.search_running = False
if "last_search" not in st.session_state:
    st.session_state.last_search = None

# ═══════════════════════════════════════════════════════════════════════════════
# LOAD CURRENT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════

def load_results():
    """Load current search results from JSON"""
    try:
        results_file = os.path.join(os.path.dirname(__file__), "data", "current_results.json")
        if os.path.exists(results_file):
            with open(results_file, "r") as f:
                return json.load(f)
    except:
        pass
    return []

def save_results(results):
    """Save current search results to JSON"""
    try:
        results_dir = os.path.join(os.path.dirname(__file__), "data")
        os.makedirs(results_dir, exist_ok=True)
        with open(os.path.join(results_dir, "current_results.json"), "w") as f:
            json.dump(results, f, indent=2)
    except:
        pass

# ═══════════════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("# 🤖 EventBot Pro")
st.markdown("**Conference & Event Discovery Engine**")
st.divider()

# ─────────────────────────────────────────────────────────────────────────────
# TOP METRICS
# ─────────────────────────────────────────────────────────────────────────────

current_results = load_results()

col1, col2, col3, col4 = st.columns(4)

with col1:
    st.markdown(f"""
    <div class="stat-box">
        <div class="stat-num">{len(current_results)}</div>
        <div class="stat-label">Events Found</div>
    </div>
    """, unsafe_allow_html=True)

with col2:
    contacts_count = len([r for r in current_results if r.get("contact_person")])
    st.markdown(f"""
    <div class="stat-box">
        <div class="stat-num">{contacts_count}</div>
        <div class="stat-label">With Contacts</div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    emails_count = len([r for r in current_results if r.get("email")])
    st.markdown(f"""
    <div class="stat-box">
        <div class="stat-num">{emails_count}</div>
        <div class="stat-label">With Emails</div>
    </div>
    """, unsafe_allow_html=True)

with col4:
    venues_count = len(set([r.get("venue_name", "") for r in current_results if r.get("venue_name")]))
    st.markdown(f"""
    <div class="stat-box">
        <div class="stat-num">{venues_count}</div>
        <div class="stat-label">Venues</div>
    </div>
    """, unsafe_allow_html=True)

st.divider()

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN INTERFACE
# ═══════════════════════════════════════════════════════════════════════════════

col_search, col_live = st.columns([1, 3])

# ─────────────────────────────────────────────────────────────────────────────
# LEFT: SEARCH PANEL
# ─────────────────────────────────────────────────────────────────────────────

with col_search:
    st.markdown("### 🔍 Search")

    # Venue selection
    venue_names = sorted(list(set([v["name"] for v in VENUES])))
    selected_venue = st.selectbox("Select Venue", venue_names, help="Pick a venue to search")

    # Year range
    st.markdown("**Year Range**")
    col_y1, col_y2 = st.columns(2)
    with col_y1:
        start_year = st.number_input("From", 2026, 2030, 2026)
    with col_y2:
        end_year = st.number_input("To", 2026, 2030, 2026)

    st.markdown("---")

    # Start button
    if st.button("▶️ START SEARCH", use_container_width=True, key="search_btn"):
        st.session_state.search_running = True
        st.session_state.current_results = []  # FRESH START - NO OLD DATA
        st.session_state.last_search = {
            "venue": selected_venue,
            "start_year": start_year,
            "end_year": end_year,
            "started": datetime.now().isoformat()
        }
        save_results([])
        st.rerun()

    st.markdown("---")

    # Filters
    st.markdown("### 🔎 Filters")

    if current_results:
        # Search by text
        search_text = st.text_input("Search events", placeholder="Event name, contact...")

        # Filter by city
        cities = list(set([r.get("city", "") for r in current_results if r.get("city")]))
        selected_cities = st.multiselect("Cities", cities, default=cities)

        # Filter by contact status
        has_contact = st.checkbox("Has contact person", value=False)
        has_email = st.checkbox("Has email", value=False)
        has_phone = st.checkbox("Has phone", value=False)

    else:
        search_text = ""
        selected_cities = []
        has_contact = False
        has_email = False
        has_phone = False

# ─────────────────────────────────────────────────────────────────────────────
# RIGHT: LIVE TRACKER & RESULTS
# ─────────────────────────────────────────────────────────────────────────────

with col_live:
    st.markdown("### 📊 Live Tracker")

    if st.session_state.search_running:
        st.markdown('<div class="status-running"><strong>⏳ Searching...</strong><br>Results appear below as they are discovered.</div>', unsafe_allow_html=True)

        # Run background search
        def run_search():
            try:
                cmd = [
                    "python", "scraper.py",
                    "--venue", selected_venue,
                    "--years", str(start_year), str(end_year)
                ]
                subprocess.run(cmd, cwd=os.path.dirname(__file__), timeout=3600)
                st.session_state.search_running = False
            except Exception as e:
                st.error(f"Error: {e}")
                st.session_state.search_running = False

        import threading
        thread = threading.Thread(target=run_search, daemon=True)
        thread.start()
        time.sleep(1)
        st.rerun()

    # Display results
    if current_results:
        # Apply filters
        filtered = current_results

        if search_text:
            filtered = [r for r in filtered if search_text.lower() in str(r).lower()]

        if selected_cities:
            filtered = [r for r in filtered if r.get("city", "") in selected_cities]

        if has_contact:
            filtered = [r for r in filtered if r.get("contact_person")]

        if has_email:
            filtered = [r for r in filtered if r.get("email")]

        if has_phone:
            filtered = [r for r in filtered if r.get("phone")]

        # Display results table
        st.markdown(f"**Found {len(filtered)} events** (Total: {len(current_results)})")

        if filtered:
            # Create DataFrame
            df_data = []
            for r in filtered:
                df_data.append({
                    "Date": r.get("event_dates", ""),
                    "Event": r.get("event_name", "")[:80],
                    "Venue": r.get("venue_name", ""),
                    "City": r.get("city", ""),
                    "Contact": r.get("contact_person", ""),
                    "Title": r.get("contact_title", ""),
                    "Email": r.get("email", ""),
                    "Phone": r.get("phone", ""),
                })

            df = pd.DataFrame(df_data)
            st.dataframe(df, use_container_width=True, height=400)

            # Quick stats
            st.markdown("---")
            col_s1, col_s2, col_s3 = st.columns(3)
            with col_s1:
                st.metric("With Contact", len([r for r in filtered if r.get("contact_person")]))
            with col_s2:
                st.metric("With Email", len([r for r in filtered if r.get("email")]))
            with col_s3:
                st.metric("With Phone", len([r for r in filtered if r.get("phone")]))

            # Export
            st.markdown("---")
            st.markdown("### 📥 Export")

            # Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = Workbook()
                ws = wb.active
                ws.title = "Events"

                headers = ["Date", "Event Name", "Venue", "City", "Contact Person", "Title", "Email", "Phone"]
                ws.append(headers)

                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)

                for r in filtered:
                    ws.append([
                        r.get("event_dates", ""),
                        r.get("event_name", ""),
                        r.get("venue_name", ""),
                        r.get("city", ""),
                        r.get("contact_person", ""),
                        r.get("contact_title", ""),
                        r.get("email", ""),
                        r.get("phone", ""),
                    ])

                for column in ws.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)

                filename = f"EventBot_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                wb.save(filename)

                with open(filename, "rb") as f:
                    st.download_button("📥 Download Excel", f, filename)

                os.remove(filename)

            except Exception as e:
                st.warning(f"Excel export: {e}")

            # CSV
            import csv
            from io import StringIO

            csv_buffer = StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow(headers)
            for r in filtered:
                writer.writerow([r.get(h.lower().replace(" ", "_"), "") for h in headers])

            st.download_button("📥 Download CSV", csv_buffer.getvalue(), f"EventBot_{datetime.now().strftime('%Y%m%d_%H%M')}.csv", "text/csv")

        else:
            st.info("No results match your filters")

    elif not st.session_state.search_running:
        if len(current_results) == 0 and st.session_state.last_search is None:
            st.info("👉 Select a venue and click START SEARCH")

st.divider()
st.caption(f"EventBot Pro v5 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
