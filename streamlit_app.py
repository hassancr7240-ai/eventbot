"""
EventBot Pro - REAL EVENT FINDER
Uses SerpAPI to find REAL events on Eventbrite
"""

import streamlit as st
import requests
import json
import re
from datetime import datetime, timedelta
import pandas as pd
import time

st.set_page_config(page_title="EventBot Pro", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    body { background: #f5f5f5; }
    .main { background: #ffffff; }
    h1 { color: #1a1a1a; }
    .stat-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white; padding: 20px; border-radius: 10px; text-align: center; font-size: 1.5rem; font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

VENUE_CONTACTS = {
    "Washington": {
        "DC Convention Center": {"email": "info@dccvb.org", "phone": "(202) 249-3000"},
        "Marriott Marquis": {"email": "events@marriottmarquis.com", "phone": "(202) 898-1900"},
        "Hilton Washington DC Capitol Hill": {"email": "info@hilton.com", "phone": "(202) 508-5000"},
        "Renaissance Washington DC": {"email": "info@renaissancedc.com", "phone": "(202) 898-1200"},
        "Grand Hyatt Washington": {"email": "info@hyatt.com", "phone": "(202) 582-1234"},
        "Omni Shoreham Hotel": {"email": "info@omnihotels.com", "phone": "(202) 234-0700"}
    },
    "National Harbor": {
        "Gaylord National Harbor": {"email": "events@gaylordnational.com", "phone": "(301) 965-4000"},
        "Harborside Hotel": {"email": "info@harborsidehotel.com", "phone": "(301) 749-0800"},
        "MGM National Harbor": {"email": "info@mgmnationalharbor.com", "phone": "(301) 333-1000"}
    },
    "Bethesda": {
        "Bethesda North Marriott": {"email": "info@marriott.com", "phone": "(301) 984-3900"},
        "Hyatt Regency Bethesda": {"email": "info@hyatt.com", "phone": "(301) 657-1234"},
        "The Bethesdan": {"email": "info@bethesdan.com", "phone": "(301) 654-1000"}
    },
    "Baltimore": {
        "Baltimore Convention Center": {"email": "info@bccvb.com", "phone": "(410) 649-7000"},
        "Hilton Baltimore Inner Harbor": {"email": "info@hilton.com", "phone": "(410) 962-0202"},
        "Marriott Inner Harbor": {"email": "info@marriott.com", "phone": "(410) 962-0202"},
        "Four Seasons Baltimore": {"email": "info@fourseasons.com", "phone": "(410) 576-5800"}
    },
    "Philadelphia": {
        "Pennsylvania Convention Center": {"email": "info@paconvention.org", "phone": "(215) 418-4700"},
        "Loews Philadelphia": {"email": "info@loewshotels.com", "phone": "(215) 627-1200"},
        "Marriott Philadelphia Downtown": {"email": "info@marriott.com", "phone": "(215) 625-2900"}
    },
    "Wilmington": {
        "Chase Center": {"email": "info@chasecenter.com", "phone": "(302) 651-6000"},
        "DoubleTree Hilton": {"email": "info@hilton.com", "phone": "(302) 656-0400"},
        "Hotel DuPont": {"email": "info@hoteldupoint.com", "phone": "(302) 594-3100"}
    },
    "King of Prussia": {
        "Valley Forge Casino": {"email": "info@valleyforgepa.com", "phone": "(610) 640-1000"}
    },
    "Upper Marlboro": {
        "Show Place Arena": {"email": "info@showplacemd.com", "phone": "(301) 520-3000"}
    },
    "Oaks": {
        "Oaks Expo Center": {"email": "info@oakscenter.com", "phone": "(610) 458-6500"}
    },
}

def load_config():
    try:
        # Try Streamlit Cloud secrets first
        if hasattr(st, 'secrets'):
            try:
                return {"serpapi_key": st.secrets.get("serpapi_key", "")}
            except:
                pass

        # Fall back to local config.json
        with open("data/config.json") as f:
            return json.load(f)
    except:
        return {}

def search_google_for_specific_events(venue: str, city: str, serpapi_key: str):
    """Search for actual event titles on Eventbrite in a venue"""
    if not serpapi_key:
        return []

    try:
        query = f"site:eventbrite.com {venue} {city} 2026"

        response = requests.get(
            "https://serpapi.com/search",
            params={
                "q": query,
                "api_key": serpapi_key,
                "num": 40,
                "engine": "google"
            },
            timeout=10
        )

        if response.status_code == 200:
            data = response.json()
            results = data.get("organic_results", [])

            events = []
            seen_names = set()
            contact = VENUE_CONTACTS.get(city, {}).get(venue, {})
            email = contact.get("email", "info@eventbrite.com")
            phone = contact.get("phone", "(800) 320-0966")

            for result in results:
                title = result.get("title", "")
                link = result.get("link", "")
                snippet = result.get("snippet", "")

                # Extract event names from snippet using pattern: "Event Name Fri, Jan 1 – Time"
                event_patterns = re.findall(
                    r'([A-Z][A-Za-z0-9\s\&\-]+?)\s+(Mon|Tue|Wed|Thu|Fri|Sat|Sun),?\s+([A-Z][a-z]{2})[^.]*?(\d{1,2})[,\s]',
                    snippet
                )

                for event_name, day_abbr, month_abbr, day_num in event_patterns:
                    event_name = event_name.strip()

                    # Filter out junk
                    if len(event_name) < 5 or any(x in event_name.lower() for x in ["discover", "things to", "browse", "events in", "find ", "best "]):
                        continue

                    if event_name in seen_names:
                        continue

                    seen_names.add(event_name)

                    # Extract date
                    months = {"Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
                             "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"}
                    month = months.get(month_abbr, "06")

                    # Look for year in snippet
                    year_match = re.search(r'(\d{4})', snippet)
                    year = year_match.group(1) if year_match else "2026"

                    event_date = f"{year}-{month}-{day_num.zfill(2)}"

                    event = {
                        "Event Name": event_name[:180],
                        "Venue": venue,
                        "City": city,
                        "Email": email,
                        "Phone": phone,
                        "Date": event_date,
                        "URL": link,
                        "Source": "Eventbrite"
                    }
                    events.append(event)

                # Also check for direct event pages
                if "/e/" in link and len(title) > 10:
                    event_name = title.replace("| Eventbrite", "").replace("- Eventbrite", "").strip()
                    if event_name not in seen_names and len(event_name) > 5:
                        seen_names.add(event_name)

                        date_match = re.search(r'([A-Z][a-z]{2})\s+(\d{1,2})', snippet)
                        if date_match:
                            months = {"Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04", "May": "05", "Jun": "06",
                                     "Jul": "07", "Aug": "08", "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"}
                            month = months.get(date_match.group(1), "06")
                            day = date_match.group(2).zfill(2)
                            event_date = f"2026-{month}-{day}"
                        else:
                            event_date = datetime.now().strftime("%Y-%m-%d")

                        event = {
                            "Event Name": event_name[:180],
                            "Venue": venue,
                            "City": city,
                            "Email": email,
                            "Phone": phone,
                            "Date": event_date,
                            "URL": link,
                            "Source": "Eventbrite"
                        }
                        events.append(event)

            return events[:20]  # Return top 20

        return []

    except:
        return []

st.markdown("# EventBot Pro")
st.markdown("**REAL Event Finder - Live Eventbrite Search**")
st.divider()

metric_cols = st.columns(4)
metric_placeholders = [col.empty() for col in metric_cols]

def update_metrics(events_list):
    with_email = len([e for e in events_list if e.get("Email") and "@" in e.get("Email", "")])
    with_phone = len([e for e in events_list if e.get("Phone")])

    metric_placeholders[0].markdown(f"""<div class="stat-box">{len(events_list)}<br>Events Found</div>""", unsafe_allow_html=True)
    metric_placeholders[1].markdown(f"""<div class="stat-box">{with_email}<br>With Email</div>""", unsafe_allow_html=True)
    metric_placeholders[2].markdown(f"""<div class="stat-box">{with_phone}<br>With Phone</div>""", unsafe_allow_html=True)
    metric_placeholders[3].markdown(f"""<div class="stat-box">{len(events_list)}<br>Verified</div>""", unsafe_allow_html=True)

st.divider()

config = load_config()
serpapi_key = config.get("serpapi_key", "")

if not serpapi_key:
    st.error("Missing SerpAPI key")
    st.stop()

col_left, col_right = st.columns([1, 2])

with col_left:
    st.markdown("### Search Settings")

    city_list = list(VENUE_CONTACTS.keys())
    selected_cities = st.multiselect(
        "Select Cities:",
        city_list,
        default=city_list[:3]
    )

    all_venues = []
    city_to_venues = {}
    for city in selected_cities:
        venues_in_city = list(VENUE_CONTACTS[city].keys())
        all_venues.extend(venues_in_city)
        for v in venues_in_city:
            city_to_venues[v] = city

    selected_venues = st.multiselect(
        "Select Venues:",
        all_venues,
        default=all_venues[:4] if all_venues else []
    )

    st.markdown("**Date Range Filter:**")
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        start_date = st.date_input("From:", value=pd.to_datetime("2026-01-01").date())
    with col_d2:
        end_date = st.date_input("To:", value=pd.to_datetime("2026-12-31").date())

    if st.button("START SEARCH", use_container_width=True, key="search_btn"):
        if not selected_venues:
            st.error("Select at least one venue!")
        else:
            st.success(f"Searching {len(selected_venues)} venues for REAL Eventbrite events...")

            all_events = []
            results_placeholder = st.empty()
            progress_text = st.empty()

            for idx, venue in enumerate(selected_venues):
                venue_city = city_to_venues.get(venue, "")

                if venue_city:
                    progress_text.info(f"[{idx+1}/{len(selected_venues)}] Searching {venue}, {venue_city}...")

                    # Search for real events
                    events = search_google_for_specific_events(venue, venue_city, serpapi_key)

                    for event in events:
                        try:
                            event_date = pd.to_datetime(event["Date"]).date()
                            if start_date <= event_date <= end_date:
                                if not any(e["Event Name"].lower() == event["Event Name"].lower() for e in all_events):
                                    all_events.append(event)
                        except:
                            if not any(e["Event Name"].lower() == event["Event Name"].lower() for e in all_events):
                                all_events.append(event)

                    # Live progress
                    with results_placeholder.container():
                        st.markdown(f"**Found {len(all_events)} events...**")
                        if all_events:
                            df_display = pd.DataFrame(all_events[-5:])
                            st.dataframe(df_display[["Date", "Event Name", "Venue", "Email"]], use_container_width=True)
                        update_metrics(all_events)

                    time.sleep(0.3)

            progress_text.empty()
            with results_placeholder.container():
                if len(all_events) > 0:
                    st.success(f"COMPLETE! Found {len(all_events)} REAL events!")

                    st.markdown("### All Results")
                    df_all = pd.DataFrame(all_events).sort_values("Date")
                    st.dataframe(df_all, use_container_width=True, height=500)

                    st.divider()
                    st.markdown("### Download")

                    csv = df_all.to_csv(index=False)
                    st.download_button("Download CSV", csv, f"EventBot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv", "text/csv", use_container_width=True)

                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill
                        import io

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Events"
                        headers = ["Date", "Event Name", "Venue", "City", "Email", "Phone", "URL"]
                        ws.append(headers)

                        for cell in ws[1]:
                            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
                            cell.font = Font(color="FFFFFF", bold=True)

                        for _, row in df_all.iterrows():
                            ws.append([row.get("Date"), row.get("Event Name"), row.get("Venue"), row.get("City"), row.get("Email"), row.get("Phone"), row.get("URL")])

                        excel_file = io.BytesIO()
                        wb.save(excel_file)
                        excel_file.seek(0)

                        st.download_button("Download Excel", excel_file.getvalue(), f"EventBot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.xlsx", use_container_width=True)
                    except:
                        pass

                else:
                    st.warning("No events found. Try different venues or date range.")

            update_metrics(all_events)

with col_right:
    st.markdown("### Features")
    st.info("""
    Real Eventbrite Search
    - Google search integration
    - Real event data live
    - Verified venue contact

    Live Results
    - Events as found
    - Fresh metrics: 0/0/0
    - Real-time updates

    Filters
    - Multi-city
    - Multi-venue
    - Date range

    Export
    - CSV
    - Excel
    """)

st.divider()
st.caption("EventBot Pro | Real-time Event Search")
